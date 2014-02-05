#!/usr/bin/env python

from cStringIO import StringIO
import datetime

from openpyxl.reader.excel import load_workbook

from csvkit import CSVKitWriter 
from csvkit.typeinference import NULL_TIME

def normalize_datetime(dt):
    if dt.microsecond == 0:
        return dt

    ms = dt.microsecond

    if ms < 1000:
        return dt.replace(microsecond=0)
    elif ms > 999000:
        return dt.replace(microsecond=0) + datetime.timedelta(seconds=1)

    return dt

def xlsx2csv(f, output=None, **kwargs):
    """
    Convert an Excel .xlsx file to csv.

    Note: Unlike other convertor's, this one allows output columns to contain mixed data types.
    Blank headers are also possible.
    """
    streaming = True if output else False

    if not streaming:
        output = StringIO()

    book = load_workbook(f, use_iterators=True)
    if 'sheet' in kwargs:
        sheet = book.get_sheet_by_name(kwargs['sheet'])
        kwargs.pop('sheet')
    else:
        sheet = book.get_active_sheet()

    writer = CSVKitWriter(output, **kwargs)

    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            writer.writerow([c.internal_value for c in row]) 
            continue

        out_row = []

        for c in row:
            value = c.internal_value

            if value.__class__ is datetime.datetime:
                if value.time() != NULL_TIME:
                    value = normalize_datetime(value)
                else:
                    value = value.date()
            elif value.__class__ is float:
                if value % 1 == 0:
                    value = int(value)

            if value.__class__ in (datetime.datetime, datetime.date, datetime.time):
                value = value.isoformat()

            out_row.append(value)

        writer.writerow(out_row)

    if not streaming:
        data = output.getvalue()
        return data

    # Return empty string when streaming
    return ''

